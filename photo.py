#this only works on JPEG images
from PIL import Image
from PIL.ExifTags import TAGS
import os
from openpyxl import Workbook
photolist = []
directory = './photos/'

for filename in os.listdir(directory):

    dict = {}

    if filename.endswith(".jpg") or filename.endswith(".JPG"):
        # read the image data using PIL
        image = Image.open(directory + filename)
        # extract EXIF data
        exifdata = image.getexif()

        for tag_id in exifdata:
            # get the tag name, instead of human unreadable tag id
            tag = TAGS.get(tag_id, tag_id)
            data = exifdata.get(tag_id)
            # decode bytes
            if isinstance(data, bytes):
                data = data.decode()
            # print(f"{tag:25}: {data}")
            dict[tag] = data

    photolist.append(dict)

print(photolist)




# print(dict)



# # path to the image or video
# imagename = "./IMG_5999.JPG"
#
# # read the image data using PIL
# image = Image.open(imagename)

# # extract EXIF data
# exifdata = image.getexif()

# #create a workbook
# wb = Workbook()
# # grab the active worksheet
# # ws = wb.active
# dict = {}

# # iterating over all EXIF data fields
# for tag_id in exifdata:
#     # get the tag name, instead of human unreadable tag id
#     tag = TAGS.get(tag_id, tag_id)
#     data = exifdata.get(tag_id)
#     # decode bytes
#     if isinstance(data, bytes):
#         data = data.decode()
#     # print(f"{tag:25}: {data}")
#     dict[tag]=data
#
# print(dict)
#
# # Data can be assigned directly to cells
# ws['A1'] = 42
#
# # Rows can also be appended
# ws.append([1, 2, 3])
#
# # Python types will automatically be converted
# import datetime
# ws['A2'] = datetime.datetime.now()
#
# # Save the file
# wb.save("sample.xlsx")